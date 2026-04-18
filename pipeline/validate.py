#!/usr/bin/env python3
"""Retrospective audit tool for Hawks Scouting game data.

Audits already-ingested games by cross-checking their markdown play logs
against the Excel/JSON repository. Reports discrepancies without modifying
any data. Uses the same cross-checks as ingest.py's pre-write validation.

Usage:
    python pipeline/validate.py games/2026-04-17_OKLN_at_RVRH.md   # single game
    python pipeline/validate.py --game-id 2026-04-17_OKLN_at_RVRH  # by ID
    python pipeline/validate.py --all                               # audit all games
    python pipeline/validate.py --since 2026-04-01 --all           # date-bounded
    python pipeline/validate.py --verbose                           # show parser detail
    python pipeline/validate.py --json                              # machine-readable

Exit codes:
  0 — all audited games pass
  1 — one or more games fail
  2 — parser error on any markdown (structural problem)
"""

import argparse
import json
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

sys.path.insert(0, str(Path(__file__).resolve().parent))
from validate_core import (
    parse_play_log, run_all_checks,
    extract_pitcher_appearances, extract_batter_appearances,
)

REPO_ROOT = Path(__file__).resolve().parent.parent
GAMES_DIR = REPO_ROOT / "games"
EXCEL_FILE = REPO_ROOT / "data" / "RiverHill_Repository_Master.xlsx"
JSON_FILE = REPO_ROOT / "public" / "repository.json"


# ─── Loaders ─────────────────────────────────────────────────────────────────

def _sheet_to_rows(ws):
    """Convert openpyxl worksheet to list of dicts."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    result = []
    for r in rows[1:]:
        if not r or all(c is None or c == "" for c in r):
            continue
        result.append({headers[i]: r[i] for i in range(len(headers))})
    return result


def load_game_from_xlsx(game_id: str) -> dict:
    """Load a single game's rows from the Excel repository.

    Returns a dict shaped like Claude's JSON output:
    { "game_log": {...}, "batting": [...], "pitching": [...], "fielding": [...] }
    """
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    data = {"game_log": {}, "batting": [], "pitching": [], "fielding": []}

    for sheet_name, key in [("Game_Log", "game_log"), ("Batting", "batting"),
                             ("Pitching", "pitching"), ("Fielding", "fielding")]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = _sheet_to_rows(ws)
        matching = [r for r in rows if str(r.get("Game_ID", "")) == game_id]
        if sheet_name == "Game_Log":
            data[key] = matching[0] if matching else {}
        else:
            data[key] = matching

    wb.close()
    return data


def resolve_markdown_path(game_id: str) -> Path:
    """Find the markdown file for a Game_ID."""
    p = GAMES_DIR / f"{game_id}.md"
    if not p.exists():
        raise FileNotFoundError(f"Markdown not found: {p}")
    return p


def game_id_from_md_path(md_path: Path) -> str:
    """Derive Game_ID from a markdown filename."""
    stem = md_path.stem
    if re.match(r'\d{4}-\d{2}-\d{2}_[A-Z]{3,5}_at_[A-Z]{3,5}', stem):
        return stem
    return stem  # fall back to raw stem (may not match any Excel row)


# ─── Audit orchestration ─────────────────────────────────────────────────────

def audit_game(md_path: Path, verbose: bool = False) -> dict:
    """Audit a single game. Returns a result dict with status + discrepancies."""
    game_id = game_id_from_md_path(md_path)

    try:
        md_text = md_path.read_text()
    except Exception as e:
        return {"game_id": game_id, "status": "parse_error", "error": f"read: {e}", "discrepancies": []}

    try:
        play_log = parse_play_log(md_text, game_id=game_id)
    except Exception as e:
        return {"game_id": game_id, "status": "parse_error", "error": f"parse: {e}", "discrepancies": []}

    if not play_log.plays:
        return {
            "game_id": game_id, "status": "parse_error",
            "error": "no play log rows found in markdown",
            "discrepancies": [],
        }

    try:
        data = load_game_from_xlsx(game_id)
    except Exception as e:
        return {"game_id": game_id, "status": "load_error", "error": f"load: {e}", "discrepancies": []}

    if not data.get("game_log"):
        return {
            "game_id": game_id, "status": "not_ingested",
            "error": "no Game_Log row in Excel for this Game_ID",
            "discrepancies": [],
        }

    report = run_all_checks(data, play_log)

    result = {
        "game_id": game_id,
        "status": "ok" if report.ok else "fail",
        "discrepancies": [
            {
                "check": d.check, "team": d.team,
                "expected": d.expected, "actual": d.actual,
                "details": d.details,
            } for d in report.discrepancies
        ],
        "summary": report.summary,
    }

    if verbose:
        # Include parser inferred segments
        pitchers = extract_pitcher_appearances(play_log)
        batters = extract_batter_appearances(play_log)
        result["parser"] = {
            "teams": {"away": play_log.away_team, "home": play_log.home_team},
            "innings_played": play_log.innings_played,
            "plays": len(play_log.plays),
            "pitchers": {
                team: {name: {"outs": t.outs, "bf": t.bf, "h": t.h, "bb": t.bb, "hbp": t.hbp, "k": t.k}
                       for name, t in segs.items()}
                for team, segs in pitchers.items() if team
            },
            "batter_counts": {
                team: {name: t.pa for name, t in segs.items()}
                for team, segs in batters.items() if team
            },
        }

    return result


def select_games(args) -> list:
    """Build the list of markdown paths to audit based on CLI args."""
    if args.game_md:
        return [Path(args.game_md)]
    if args.game_id:
        return [resolve_markdown_path(args.game_id)]
    if args.all:
        paths = sorted(GAMES_DIR.glob("*.md"))
        # Skip UNKNOWN_ prefixed files — they aren't ingested by design
        paths = [p for p in paths if not p.stem.startswith("UNKNOWN_")]
        if args.since:
            paths = [p for p in paths if p.stem >= args.since]
        if args.until:
            paths = [p for p in paths if p.stem <= args.until]
        return paths
    return []


# ─── Output ──────────────────────────────────────────────────────────────────

def format_human(results, verbose=False):
    """Human-readable report."""
    lines = []
    ok_count = 0
    fail_count = 0
    error_count = 0
    fail_ids = []

    for r in results:
        gid = r["game_id"]
        status = r["status"]

        if status == "ok":
            ok_count += 1
            if verbose:
                lines.append(f"[OK]   {gid}")
            continue

        if status == "fail":
            fail_count += 1
            fail_ids.append(gid)
            lines.append(f"[FAIL] {gid}")
            for d in r["discrepancies"]:
                lines.append(f"  {d['check']} {d['team']}: expected {d['expected']}, got {d['actual']}")
                lines.append(f"    {d['details']}")
            if verbose and "parser" in r:
                lines.append(f"  Parser: {json.dumps(r['parser'], indent=2)}")
        elif status in ("parse_error", "load_error"):
            error_count += 1
            lines.append(f"[ERR]  {gid}: {r.get('error', 'unknown error')}")
        elif status == "not_ingested":
            # Not an error — game markdown exists but wasn't ingested (common for UNKNOWN_ or gate-failed games)
            if verbose:
                lines.append(f"[SKIP] {gid}: not ingested")

    summary = f"\n{len(results)} games audited: {ok_count} OK, {fail_count} failing, {error_count} errors"
    if fail_ids:
        summary += f"\nFailing Game_IDs:\n  " + "\n  ".join(fail_ids)
    lines.append(summary)

    return "\n".join(lines)


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Audit Hawks Scouting game data by cross-checking markdown vs repository",
    )
    parser.add_argument("game_md", nargs="?", help="Path to a game markdown file")
    parser.add_argument("--game-id", help="Game_ID (e.g., 2026-04-17_OKLN_at_RVRH)")
    parser.add_argument("--all", action="store_true", help="Audit all games in games/")
    parser.add_argument("--since", help="Only games with Game_ID >= this (YYYY-MM-DD)")
    parser.add_argument("--until", help="Only games with Game_ID <= this (YYYY-MM-DD)")
    parser.add_argument("--verbose", action="store_true", help="Show parser inferred segments and OK games")
    parser.add_argument("--json", action="store_true", help="Output machine-readable JSON")
    args = parser.parse_args()

    paths = select_games(args)
    if not paths:
        sys.exit("Nothing to audit. Specify a file, --game-id, or --all.")

    results = []
    parse_errors = 0
    fail_count = 0

    for i, p in enumerate(paths, 1):
        if len(paths) > 10 and not args.json:
            print(f"  [{i}/{len(paths)}] {p.stem}", end="\r", file=sys.stderr, flush=True)
        result = audit_game(p, verbose=args.verbose)
        results.append(result)
        if result["status"] == "parse_error":
            parse_errors += 1
        elif result["status"] == "fail":
            fail_count += 1

    if len(paths) > 10 and not args.json:
        print(" " * 80, end="\r", file=sys.stderr)  # clear progress

    if args.json:
        print(json.dumps(results, indent=2))
    else:
        print(format_human(results, verbose=args.verbose))

    # Exit code
    if parse_errors > 0:
        sys.exit(2)
    if fail_count > 0:
        sys.exit(1)
    sys.exit(0)


if __name__ == "__main__":
    main()
