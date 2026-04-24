"""
canonicalize_aliases.py

One-shot rewrite to collapse legacy alias codes into their canonical
equivalents across the entire repository.

Aliases handled (legacy → canonical):
    MT.H → MTHB
    ST.M → STMC
    KNTS → KTIS
    CNTR → CNTY
    CML  → CMLW

Touches:
    - data/RiverHill_Repository_Master.xlsx
        Game_Log:                 Game_ID, Focal_Team, Away_Team, Home_Team
        Batting/Pitching/Fielding: Game_ID, Team, Opponent
        Roster:                    Team_Code
    - games/*.md filenames where AWAY or HOME slot is a legacy code
    - public/games/*.md filenames (same)

Safe to run repeatedly. Creates a timestamped Excel backup before writing.

Note: this script does not edit markdown *content* — play logs that
mention these codes in body text are left as-is. The .md filename and
all Excel columns are the operational join keys; body text is human
reference.

Usage:
    python pipeline/canonicalize_aliases.py            # dry run
    python pipeline/canonicalize_aliases.py --apply    # write changes
"""

import argparse
import re
import shutil
from datetime import datetime
from pathlib import Path
from collections import Counter

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl not found.  Run: pip install openpyxl")

# ─── Paths ────────────────────────────────────────────────────────────────────

REPO_ROOT       = Path(__file__).resolve().parent.parent
EXCEL_PATH      = REPO_ROOT / "data" / "RiverHill_Repository_Master.xlsx"
BACKUP_DIR      = REPO_ROOT / "data" / "backups"
GAMES_DIR       = REPO_ROOT / "games"
PUBLIC_GAMES    = REPO_ROOT / "public" / "games"

# ─── Alias map ────────────────────────────────────────────────────────────────

ALIASES = {
    "MT.H": "MTHB",
    "ST.M": "STMC",
    "KNTS": "KTIS",
    "CNTR": "CNTY",
    "CML":  "CMLW",
}

# Sheet → list of (column_name, kind)
# kind = "code"      → exact match in ALIASES → replace
#        "game_id"   → parse YYYY-MM-DD_AWAY_at_HOME and substitute parts
SHEET_COLUMNS = {
    "Game_Log": [
        ("Game_ID",    "game_id"),
        ("Focal_Team", "code"),
        ("Away_Team",  "code"),
        ("Home_Team",  "code"),
    ],
    "Batting": [
        ("Game_ID",  "game_id"),
        ("Team",     "code"),
        ("Opponent", "code"),
    ],
    "Pitching": [
        ("Game_ID",  "game_id"),
        ("Team",     "code"),
        ("Opponent", "code"),
    ],
    "Fielding": [
        ("Game_ID",  "game_id"),
        ("Team",     "code"),
        ("Opponent", "code"),
    ],
    "Roster": [
        ("Team_Code", "code"),
    ],
}

GAME_ID_RE = re.compile(r"^(\d{4}-\d{2}-\d{2})_(.+?)_at_(.+)$")

# ─── Helpers ──────────────────────────────────────────────────────────────────

def canonicalize_code(val):
    """Exact-match alias substitution for a single code value."""
    if not isinstance(val, str):
        return val, False
    stripped = val.strip()
    if stripped in ALIASES:
        return ALIASES[stripped], True
    return val, False


def canonicalize_game_id(val):
    """Replace alias codes inside the AWAY/HOME slots of a Game_ID string."""
    if not isinstance(val, str):
        return val, False
    m = GAME_ID_RE.match(val.strip())
    if not m:
        return val, False
    date, away, home = m.groups()
    new_away = ALIASES.get(away, away)
    new_home = ALIASES.get(home, home)
    if new_away == away and new_home == home:
        return val, False
    return f"{date}_{new_away}_at_{new_home}", True


# ─── Excel rewrite ────────────────────────────────────────────────────────────

def rewrite_excel(wb, apply_changes):
    """Walk all configured sheets/columns and rewrite alias values.
    Returns dict: (sheet, column) → count of changed cells."""
    counts = Counter()

    for sheet_name, cols in SHEET_COLUMNS.items():
        if sheet_name not in wb.sheetnames:
            print(f"WARNING: sheet '{sheet_name}' not found, skipping")
            continue

        ws      = wb[sheet_name]
        headers = [c.value for c in ws[1]]

        for col_name, kind in cols:
            try:
                col_idx = headers.index(col_name) + 1
            except ValueError:
                print(f"WARNING: column '{col_name}' not found in {sheet_name}, skipping")
                continue

            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if kind == "code":
                    new_val, changed = canonicalize_code(cell.value)
                else:
                    new_val, changed = canonicalize_game_id(cell.value)
                if changed:
                    counts[(sheet_name, col_name)] += 1
                    if apply_changes:
                        cell.value = new_val

    return counts


# ─── Filename rewrite ─────────────────────────────────────────────────────────

def collect_filename_renames(directory):
    """Return list of (old_path, new_path) for .md files whose Game_ID-style
    stem contains an alias code in the AWAY or HOME slot."""
    renames = []
    if not directory.exists():
        return renames
    for path in directory.glob("*.md"):
        new_stem, changed = canonicalize_game_id(path.stem)
        if changed:
            new_path = path.with_name(new_stem + ".md")
            renames.append((path, new_path))
    return renames


def apply_renames(renames):
    for old, new in renames:
        if new.exists() and old.resolve() != new.resolve():
            print(f"WARNING: target exists, skipping rename: {new.name}")
            continue
        old.rename(new)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true",
                        help="Write Excel + rename files (default: dry run)")
    args = parser.parse_args()

    if not EXCEL_PATH.exists():
        raise SystemExit(f"Excel not found: {EXCEL_PATH}")

    print("Aliases:")
    for k, v in ALIASES.items():
        print(f"  {k:6s} → {v}")
    print()

    # Backup before any write
    if args.apply:
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = BACKUP_DIR / f"RiverHill_Repository_Master_{ts}_pre_canonicalize.xlsx"
        shutil.copy2(EXCEL_PATH, backup)
        print(f"Backup: {backup}\n")

    # ── Excel ────────────────────────────────────────────────────────────────
    wb     = openpyxl.load_workbook(EXCEL_PATH)
    counts = rewrite_excel(wb, apply_changes=args.apply)

    print("Excel changes by sheet/column:")
    if not counts:
        print("  (none)")
    else:
        for (sheet, col), n in sorted(counts.items()):
            print(f"  {sheet:10s}  {col:12s}  {n:5d}")
    print(f"  Total: {sum(counts.values())}\n")

    if args.apply and counts:
        wb.save(EXCEL_PATH)
        print(f"Excel saved: {EXCEL_PATH}\n")

    # ── Filenames ────────────────────────────────────────────────────────────
    games_renames  = collect_filename_renames(GAMES_DIR)
    public_renames = collect_filename_renames(PUBLIC_GAMES)

    print("File renames:")
    print(f"  games/         {len(games_renames):3d}")
    print(f"  public/games/  {len(public_renames):3d}")
    if games_renames:
        print("  examples:")
        for old, new in games_renames[:5]:
            print(f"    {old.name}  →  {new.name}")
        if len(games_renames) > 5:
            print(f"    … ({len(games_renames) - 5} more)")

    if args.apply:
        apply_renames(games_renames)
        apply_renames(public_renames)
        if games_renames or public_renames:
            print("\nFiles renamed.")

    if not args.apply:
        print("\nDRY RUN — no changes written. Re-run with --apply to write.")
    else:
        total_changes = sum(counts.values()) + len(games_renames) + len(public_renames)
        if total_changes:
            print("\nNext: run pipeline/export.py to refresh public/repository.json.")
        else:
            print("\nNo changes needed.")


if __name__ == "__main__":
    main()
