#!/usr/bin/env python3
"""Ingest a transcribed game markdown into the master Excel repository via Claude API.

Architecture:
  - Python reads the Excel to check for duplicates and get existing Game_IDs
  - Claude API receives the ingest prompt + game markdown + focal team + existing Game_IDs
  - Claude returns structured JSON with rows to append to each sheet
  - Python validates gate results, appends rows to Excel, and calls export.py

Usage:
  export ANTHROPIC_API_KEY="sk-..."
  python pipeline/ingest.py games/2026-04-08_GLNL_at_GLFR.md
"""

import json
import os
import re
import subprocess
import sys
import time
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent.parent / ".env", override=True)
except ImportError:
    pass

try:
    import anthropic
except ImportError:
    sys.exit("anthropic is required: pip install anthropic")

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

REPO_ROOT = Path(__file__).resolve().parent.parent
PROMPTS_DIR = REPO_ROOT / "prompts"
PIPELINE_DIR = REPO_ROOT / "pipeline"
DATA_DIR = REPO_ROOT / "data"
EXCEL_FILE = DATA_DIR / "RiverHill_Repository_Master.xlsx"
CONFIG_FILE = PIPELINE_DIR / "config.json"

MODEL = "claude-sonnet-4-20250514"
MAX_TOKENS = 16000


# Pipeline notes audit layer (never affects write/gate behavior; I/O failures
# are swallowed inside log_flag.log_flag).
try:
    import log_flag as _log_flag_mod
except ImportError:
    _log_flag_mod = None


_GAME_ID_RE = re.compile(r'(\d{4}-\d{2}-\d{2})_([A-Z]{3,5})_at_([A-Z]{3,5})')


def _log_note(filename, ingest_result, flag_type=None, severity=None,
              team=None, detail=""):
    """Wrapper around log_flag.log_flag(): derives game_id/date/teams from the
    filename stem and records one flag (or a bare entry if flag_type is None).
    Never raises. Does nothing if log_flag isn't importable or filename doesn't
    match the Game_ID pattern.
    """
    if _log_flag_mod is None or not filename:
        return
    m = _GAME_ID_RE.match(filename)
    if not m:
        return
    flags = []
    if flag_type:
        flags.append({
            "flag_type": flag_type,
            "severity":  severity,  # None → log_flag applies DEFAULT_SEVERITY
            "team":      team,
            "detail":    detail,
        })
    try:
        _log_flag_mod.log_flag(
            game_id=filename,
            game_date=m.group(1),
            teams=[m.group(2), m.group(3)],
            ingest_result=ingest_result,
            flags=flags,
        )
    except Exception:
        pass


def load_prompt():
    """Load prompts/ingest.md verbatim."""
    path = PROMPTS_DIR / "ingest.md"
    if not path.exists():
        sys.exit(f"Ingestion prompt not found: {path}")
    return path.read_text()


def load_config():
    if not CONFIG_FILE.exists():
        sys.exit(f"Config not found: {CONFIG_FILE}")
    with open(CONFIG_FILE) as f:
        return json.load(f)


def get_existing_game_ids():
    """Read all Game_IDs from the Game_Log sheet."""
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    ws = wb["Game_Log"]
    game_ids = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row and row[0]:
            game_ids.append(str(row[0]))
    wb.close()
    return game_ids


def load_team_registry(config):
    """Build a flat list of (pattern, canonical_code) pairs from config.json.

    Patterns are ordered longest-first so 'north harford' matches before 'north'.
    Focal teams are checked before known opponents.
    """
    registry = []
    for team in config.get("focal_teams", []):
        for pattern in team.get("name_patterns", []):
            registry.append((pattern.lower(), team["code"]))
    for team in config.get("known_opponents", []):
        for pattern in team.get("name_patterns", []):
            registry.append((pattern.lower(), team["code"]))
    # Sort longest patterns first so "north harford" matches before "north"
    registry.sort(key=lambda x: -len(x[0]))
    return registry


def resolve_team_name(full_name, registry):
    """Return canonical code for a full team name, or None if unknown."""
    name_lower = full_name.lower().strip()
    for pattern, code in registry:
        if pattern in name_lower:
            return code
    return None


def build_game_code_map(game_md_text, registry, filename=None):
    """Parse the markdown header for full team names and raw codes.
    Return a dict mapping raw_code → canonical_code for this game.

    Uses name-based resolution from the team registry. Falls back to
    raw code if the team name resolves to the same code. Warns loudly
    if a team name can't be resolved.
    """
    header = game_md_text[:3000]
    if filename:
        header = filename + "\n" + header

    code_map = {}

    # Extract raw team codes from Game_ID pattern or line score
    raw_codes = []
    m = re.search(r'\d{4}-\d{2}-\d{2}_([A-Z]{3,5})_at_([A-Z]{3,5})', header)
    if m:
        raw_codes = [m.group(1), m.group(2)]
    if not raw_codes:
        m = re.search(r'Final\s+[Ss]core.*?([A-Z]{3,5})\s+\d+\s*[-–,]\s*([A-Z]{3,5})\s+\d+', header)
        if m:
            raw_codes = [m.group(1), m.group(2)]
    if not raw_codes:
        found = re.findall(r'\|\s*([A-Z]{3,5})\s*\|', header)
        if len(found) >= 2:
            raw_codes = [found[0], found[1]]

    # Extract full team names from the Teams line
    teams_match = re.search(r'Teams[:\*]*\s*(.+)', header)
    team_names = []
    if teams_match:
        line = teams_match.group(1).strip()
        parts = re.split(r'\s*(?:vs\.?|/|@)\s*', line, maxsplit=1)
        if len(parts) == 2:
            for part in parts:
                # Clean labels and code suffixes
                name = re.sub(r'\s*\(?(away|home)\)?\s*$', '', part.strip(), flags=re.I)
                name = re.sub(r'\s*\(?[A-Z]{3,5}\)?\s*$', '', name).strip()
                name = re.sub(r'^[A-Z]{3,5}\s+', '', name).strip()  # leading code
                team_names.append(name)

    # Resolve names to canonical codes
    for i, raw_code in enumerate(raw_codes):
        # Try name-based resolution first
        if i < len(team_names) and team_names[i]:
            canonical = resolve_team_name(team_names[i], registry)
            if canonical:
                if raw_code != canonical:
                    print(f"  [REGISTRY] {raw_code} → {canonical} (from name: '{team_names[i]}')")
                code_map[raw_code] = canonical
                continue

        # Try resolving the raw code itself as a name (handles cases like RVRH)
        canonical = resolve_team_name(raw_code, registry)
        if canonical:
            if raw_code != canonical:
                print(f"  [REGISTRY] {raw_code} → {canonical} (code matched as name)")
            code_map[raw_code] = canonical
        else:
            # Unknown team — pass through with warning
            print(f"  [REGISTRY] WARNING: Could not resolve code '{raw_code}'"
                  f"{' (name: ' + repr(team_names[i]) + ')' if i < len(team_names) else ''}. "
                  f"Add to config.json known_opponents to fix.")
            _name_hint = team_names[i] if i < len(team_names) else ""
            _log_note(
                filename,
                ingest_result="success",  # may be overwritten by later call; the team was preserved
                flag_type=_log_flag_mod.REGISTRY_UNKNOWN_TEAM if _log_flag_mod else None,
                team=raw_code,
                detail=(f"Unresolved team code '{raw_code}'"
                        + (f" (name hint: {_name_hint!r})" if _name_hint else "")
                        + " — raw code preserved. Add to config.json known_opponents."),
            )
            code_map[raw_code] = raw_code

    return code_map


def determine_focal_team(game_md_text, config, filename=None, registry=None):
    """Determine the focal team using the name-based team registry.

    Resolves raw team codes from the markdown header via name matching,
    then checks which resolved code is a focal team.
    """
    focal_codes = {t["code"] for t in config["focal_teams"]}
    primary_code = next((t["code"] for t in config["focal_teams"] if t.get("primary")), "RVRH")

    if registry is None:
        registry = load_team_registry(config)

    code_map = build_game_code_map(game_md_text, registry, filename)
    resolved_teams = set(code_map.values())

    # Intersect with focal teams
    found_focal = list(resolved_teams & focal_codes)

    if not found_focal:
        # Fallback: search header for focal codes directly
        header = game_md_text[:3000]
        if filename:
            header = filename + "\n" + header
        found_focal = [code for code in focal_codes
                       if re.search(r'\b' + code + r'\b', header)]

    if not found_focal:
        sys.exit("Could not determine focal team — no focal team codes found in game header. "
                 "Check config.json team registry.")

    return primary_code if primary_code in found_focal else found_focal[0]


def call_claude(system_prompt, game_md_text, focal_team, existing_game_ids):
    """Call Claude API to extract structured data from the game markdown."""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        sys.exit("Set ANTHROPIC_API_KEY environment variable")

    # Build the user message with context Claude needs
    user_message = (
        f"Focal team: {focal_team}\n\n"
        f"Existing Game_IDs in repository (for duplicate check):\n"
        f"{json.dumps(existing_game_ids)}\n\n"
        f"IMPORTANT: You cannot read or write the Excel file directly. Instead, after completing "
        f"your full analysis per the ingestion prompt (Steps 0-4, all verification gates), output "
        f"your results as a JSON block that I will use to append rows to the Excel file.\n\n"
        f"After your confirmation block, output a JSON block fenced with ```json and ``` containing:\n"
        f'{{"game_log": {{...single row as dict...}},\n'
        f' "batting": [{{...row dicts...}}],\n'
        f' "pitching": [{{...row dicts...}}],\n'
        f' "fielding": [{{...row dicts...}}],\n'
        f' "gates": {{"G1": {{"pass": true/false, "value": N}}, ...}},\n'
        f' "duplicate": false,\n'
        f' "notes": "..."}}\n\n'
        f"Use the exact column names from the locked schema. All numeric fields must be numbers, not strings.\n\n"
        f"--- GAME LOG MARKDOWN ---\n\n"
        f"{game_md_text}"
    )

    client = anthropic.Anthropic(api_key=api_key)

    print(f"Calling Claude API ({MODEL})...")
    print(f"  System prompt: {len(system_prompt):,} chars")
    print(f"  User message: {len(user_message):,} chars")

    response = client.messages.create(
        model=MODEL,
        max_tokens=MAX_TOKENS,
        system=system_prompt,
        messages=[{"role": "user", "content": user_message}],
    )

    text = response.content[0].text
    print(f"  Response: {len(text):,} chars, {response.usage.input_tokens} input / {response.usage.output_tokens} output tokens")

    # Rate-limit courtesy delay between API calls
    print("  Waiting 15s (rate-limit delay)...")
    time.sleep(15)

    return text


def parse_json_from_response(response_text):
    """Extract the JSON block from the API response."""
    # Look for ```json ... ``` block
    m = re.search(r'```json\s*\n(.*?)\n```', response_text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError as e:
            print(f"JSON parse error: {e}")
            print(f"Raw JSON block:\n{m.group(1)[:500]}")
            return None

    # Try to find raw JSON object
    m = re.search(r'\{[\s\S]*"game_log"[\s\S]*"batting"[\s\S]*\}', response_text)
    if m:
        try:
            return json.loads(m.group(0))
        except json.JSONDecodeError:
            pass

    return None


def canonicalize_player_names(data):
    """Rewrite Player/Pitcher fields in `data` to canonical full-name forms
    where possible, before appending to Excel.

    Different scorers (e.g., the focal team's vs the opponent's) often log
    the same player under different forms — short ("M Kirk") vs full
    ("Miles Kirk") — across games. Without canonicalization at ingest time,
    each form becomes a separate entry in the database and the app shows
    duplicate player rows. This function builds a per-team
    (last_name, first_initial) → canonical-full-name map from the union of
    existing Excel rows and this game's new rows, then rewrites short forms
    to the matching full where the bucket has exactly one full-name form.

    Ambiguous buckets (2+ distinct full names sharing initial+last, e.g.
    Kamden Twele + Kyle Twele both with initial K) are left UNTOUCHED — we
    can't auto-disambiguate without per-game context.

    Returns the number of new-row rewrites applied. Does not modify existing
    Excel rows; a separate backfill is required for that.
    """
    from collections import defaultdict

    def split(name):
        if not name or name == "Unknown Player":
            return None
        parts = name.strip().split()
        if len(parts) < 2:
            return None
        first = parts[0].rstrip(".")
        if not first:
            return None
        return first, " ".join(parts[1:])

    # Collect (Team, last, initial) → set of observed names from Excel + new data
    buckets = defaultdict(set)

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    for sheet_name, name_field in (("Batting", "Player"), ("Pitching", "Pitcher"), ("Fielding", "Player")):
        ws = wb[sheet_name]
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = list(row)
                team_idx = headers.index("Team")
                name_idx = headers.index(name_field)
                continue
            team = row[team_idx]
            nm = row[name_idx]
            s = split(nm)
            if not s or not team:
                continue
            first, last = s
            buckets[(team, last, first[0].upper())].add(nm)
    wb.close()

    # Add this game's rows to the buckets so a new full-form name in this
    # game's data canonicalizes its own short-form siblings even if no prior
    # game had the full form.
    for sheet_data, name_field in ((data.get("batting", []), "Player"),
                                   (data.get("pitching", []), "Pitcher"),
                                   (data.get("fielding", []), "Player")):
        for r in sheet_data:
            s = split(r.get(name_field))
            if not s:
                continue
            first, last = s
            team = r.get("Team")
            if not team:
                continue
            buckets[(team, last, first[0].upper())].add(r[name_field])

    # Build (Team, observed_name) → canonical_full rename map
    rename = {}
    for (team, last, init), names in buckets.items():
        fulls = {n for n in names if len(n.strip().split()[0].rstrip(".")) >= 2}
        shorts = {n for n in names if len(n.strip().split()[0].rstrip(".")) == 1}
        if len(fulls) >= 2 or not fulls or not shorts:
            continue  # ambiguous, all-shorts, or all-fulls — nothing to do
        canon = next(iter(fulls))
        for s in shorts:
            rename[(team, s)] = canon

    # Apply to new rows in `data`
    n = 0
    for sheet_data, name_field in ((data.get("batting", []), "Player"),
                                   (data.get("pitching", []), "Pitcher"),
                                   (data.get("fielding", []), "Player")):
        for r in sheet_data:
            key = (r.get("Team"), r.get(name_field))
            if key in rename:
                r[name_field] = rename[key]
                n += 1
    return n


def append_to_excel(data):
    """Append the extracted rows to the Excel file."""
    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Sheet column orders (must match locked schema exactly)
    SCHEMAS = {
        "Game_Log": ["Game_ID", "Game_Date", "Game_Type", "Focal_Team", "Away_Team", "Home_Team",
                     "Innings_Played", "Source_File", "Away_R", "Away_H", "Away_E",
                     "Home_R", "Home_H", "Home_E", "QA_Flag_Count", "Notes"],
        "Batting": ["Game_ID", "Game_Date", "Opponent", "Team", "Player", "PA", "AB", "H",
                    "1B", "2B", "3B", "HR", "BB", "HBP", "K", "K_L", "K_S",
                    "R", "RBI", "SB", "CS", "GDP", "SAC", "FC", "Notes"],
        "Pitching": ["Game_ID", "Game_Date", "Opponent", "Team", "Pitcher", "Outs_Recorded",
                     "BF", "H_Allowed", "1B_Allowed", "2B_Allowed", "3B_Allowed", "HR_Allowed",
                     "BB_Allowed", "HBP_Allowed", "K", "R_Allowed", "WP", "Notes"],
        "Fielding": ["Game_ID", "Game_Date", "Opponent", "Team", "Player", "Inning",
                     "Play_Ref", "Notes"],
    }

    counts = {}

    # Append Game_Log row
    ws = wb["Game_Log"]
    row_data = data["game_log"]
    ws.append([row_data.get(col, "") for col in SCHEMAS["Game_Log"]])
    counts["Game_Log"] = 1

    # Append Batting rows
    ws = wb["Batting"]
    for row_data in data["batting"]:
        ws.append([row_data.get(col, 0 if col not in ("Game_ID", "Game_Date", "Opponent", "Team", "Player", "Notes") else "") for col in SCHEMAS["Batting"]])
    counts["Batting"] = len(data["batting"])

    # Append Pitching rows
    ws = wb["Pitching"]
    for row_data in data["pitching"]:
        ws.append([row_data.get(col, 0 if col not in ("Game_ID", "Game_Date", "Opponent", "Team", "Pitcher", "Notes") else "") for col in SCHEMAS["Pitching"]])
    counts["Pitching"] = len(data["pitching"])

    # Append Fielding rows
    ws = wb["Fielding"]
    for row_data in data["fielding"]:
        ws.append([row_data.get(col, "") for col in SCHEMAS["Fielding"]])
    counts["Fielding"] = len(data["fielding"])

    wb.save(EXCEL_FILE)
    wb.close()

    return counts


def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python pipeline/ingest.py <game_markdown_file>")

    # Simple arg parsing: [--skip-crosschecks] <game_md_path>
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    skip_crosschecks = "--skip-crosschecks" in sys.argv
    if not args:
        sys.exit("Usage: python pipeline/ingest.py [--skip-crosschecks] <game_md_path>")
    game_md_path = Path(args[0])
    if not game_md_path.exists():
        sys.exit(f"File not found: {game_md_path}")

    if not EXCEL_FILE.exists():
        sys.exit(f"Excel file not found: {EXCEL_FILE}")

    system_prompt = load_prompt()
    config = load_config()
    game_md_text = game_md_path.read_text()

    # Determine focal team
    focal_team = determine_focal_team(game_md_text, config, filename=game_md_path.stem)
    print(f"Ingesting: {game_md_path.name}")
    print(f"  Focal team: {focal_team}")

    # Check for duplicates in Python first (fast check)
    existing_ids = get_existing_game_ids()
    # Extract Game_ID from the markdown to check
    gid_match = re.search(r'(\d{4}-\d{2}-\d{2})_([A-Z]{3,5})_at_([A-Z]{3,5})', game_md_path.stem)
    if gid_match:
        candidate_id = game_md_path.stem
        if candidate_id in existing_ids:
            print(f"\nDUPLICATE DETECTED: {candidate_id} already exists in the repository.")
            print("No data written — duplicate guard fired correctly.")
            _log_note(game_md_path.stem, ingest_result="skipped",
                      detail="Duplicate guard (Python-level): game already in Game_Log.")
            sys.exit(0)

    # Call Claude API
    response_text = call_claude(system_prompt, game_md_text, focal_team, existing_ids)

    # Check for duplicate detection in response
    if re.search(r"DUPLICATE\s+DETECTED", response_text, re.IGNORECASE):
        m = re.search(r"DUPLICATE\s+DETECTED.*?(?:\n|$)", response_text, re.IGNORECASE)
        print(f"\n{m.group(0).strip() if m else 'Duplicate detected by Claude'}")
        print("No data written — duplicate guard fired correctly.")
        _log_note(game_md_path.stem, ingest_result="skipped",
                  detail="Duplicate guard (Claude-level): DUPLICATE DETECTED in API response.")
        sys.exit(0)

    # Parse JSON from response
    data = parse_json_from_response(response_text)
    if not data:
        print("\nERROR: Could not parse JSON from API response.")
        print("\n--- Full API Response ---")
        print(response_text)
        sys.exit(1)

    # Apply name-based team code remapping using the registry
    # This replaces the old CODE_ALIASES system — resolves codes per-game via team names
    registry = load_team_registry(config)
    code_map = build_game_code_map(game_md_text, registry, filename=game_md_path.stem)
    TEAM_FIELDS = ("Team", "Away_Team", "Home_Team", "Focal_Team", "Opponent")

    def remap_row(row):
        for field in TEAM_FIELDS:
            if field in row and row[field] in code_map and row[field] != code_map[row[field]]:
                old = row[field]
                row[field] = code_map[old]
                print(f"  [REGISTRY] remap: {field} {old} → {row[field]}")
        return row

    if "game_log" in data:
        data["game_log"] = remap_row(data["game_log"])
    for sheet in ("batting", "pitching", "fielding"):
        if sheet in data:
            data[sheet] = [remap_row(r) for r in data[sheet]]

    # Check gates (Claude's self-reported tallies)
    gates = data.get("gates", {})
    all_passed = True
    _failed_gates = []
    for gate_name in ["G1", "G2", "G3", "G4", "G5", "G6"]:
        gate = gates.get(gate_name, {})
        passed = gate.get("pass", False)
        value = gate.get("value", "?")
        status = "✅" if passed else "❌"
        print(f"  {gate_name} {status} {value}")
        if not passed:
            all_passed = False
            _failed_gates.append((gate_name, value))

    if not all_passed:
        print("\nGATE FAILURE — not writing to Excel.")
        print(f"Notes: {data.get('notes', '')}")
        # Audit: one flag per failed gate; single ingest_result="gate_failure" entry
        if _log_flag_mod is not None:
            _flags = [{
                "flag_type": _log_flag_mod.GATE_FAILURE,
                "severity":  "hard",
                "team":      None,
                "detail":    f"{gn} failed (value={val}). Notes: {data.get('notes', '')}",
            } for gn, val in _failed_gates]
            m = _GAME_ID_RE.match(game_md_path.stem)
            if m:
                try:
                    _log_flag_mod.log_flag(
                        game_id=game_md_path.stem,
                        game_date=m.group(1),
                        teams=[m.group(2), m.group(3)],
                        ingest_result="gate_failure",
                        flags=_flags,
                    )
                except Exception:
                    pass
        sys.exit(1)

    # Python-side cross-checks (PC1-PC5): compare Claude's output JSON against
    # play-log ground truth. Catches cases where gates pass but output is missing
    # rows (e.g., missing pitcher). See pipeline/validate_core.py for details.
    if skip_crosschecks:
        print("  (cross-checks skipped via --skip-crosschecks flag)")
    else:
        try:
            from validate_core import parse_play_log, run_all_checks
            play_log = parse_play_log(game_md_text, game_id=game_md_path.stem)
            if play_log.plays:
                report = run_all_checks(data, play_log)
                for d in report.discrepancies:
                    print(f"  {d.check} ❌ {d.team}: expected {d.expected}, got {d.actual}")
                    print(f"    {d.details}")
                if not report.ok:
                    print("\nCROSS-CHECK FAILURE — not writing to Excel.")
                    print("Review discrepancies above. If the failures are legitimate")
                    print("(e.g., jersey-only team, known hit-count discrepancy), re-run with")
                    print("--skip-crosschecks to bypass. Otherwise, fix the markdown or re-ingest.")
                    # Audit: one GATE_FAILURE flag per discrepancy. Cross-check
                    # hard-fails are surfaced here (validate_core.py doesn't log
                    # them itself — per spec).
                    if _log_flag_mod is not None:
                        _cc_flags = [{
                            "flag_type": _log_flag_mod.GATE_FAILURE,
                            "severity":  "hard",
                            "team":      d.team,
                            "detail":    f"{d.check} {d.team}: expected {d.expected}, got {d.actual} ({d.details})",
                        } for d in report.discrepancies]
                        m = _GAME_ID_RE.match(game_md_path.stem)
                        if m:
                            try:
                                _log_flag_mod.log_flag(
                                    game_id=game_md_path.stem,
                                    game_date=m.group(1),
                                    teams=[m.group(2), m.group(3)],
                                    ingest_result="gate_failure",
                                    flags=_cc_flags,
                                )
                            except Exception:
                                pass
                    sys.exit(1)
                print("  PC1-PC5 ✅ all cross-checks passed")
                # Audit: forward any soft events (PC2 tolerance / Section 5
                # match / phantom pitcher skip) from validate_core to the
                # pipeline notes. These do NOT block writes.
                if _log_flag_mod is not None and report.soft_events:
                    _soft_flags = []
                    for ev in report.soft_events:
                        flag_type = getattr(_log_flag_mod, ev.kind, None)
                        if flag_type is None:
                            continue
                        _soft_flags.append({
                            "flag_type": flag_type,
                            "severity":  None,  # use DEFAULT_SEVERITY
                            "team":      ev.team,
                            "detail":    ev.detail,
                        })
                    if _soft_flags:
                        m = _GAME_ID_RE.match(game_md_path.stem)
                        if m:
                            try:
                                _log_flag_mod.log_flag(
                                    game_id=game_md_path.stem,
                                    game_date=m.group(1),
                                    teams=[m.group(2), m.group(3)],
                                    ingest_result="success",
                                    flags=_soft_flags,
                                )
                            except Exception:
                                pass
            else:
                print("  (cross-checks skipped — no play log rows parsed)")
        except ImportError:
            print("  (cross-checks skipped — validate_core.py not found)")
        except Exception as e:
            print(f"  (cross-checks skipped — error: {e})")

    if data.get("duplicate", False):
        print("\nDuplicate flag set in response — not writing.")
        sys.exit(0)

    # Canonicalize player names against the existing Excel state so this
    # game's rows merge with prior games' stats. Different scorers use
    # different forms ("M Kirk" vs "Miles Kirk") for the same player; without
    # this, the app shows duplicate per-player rows.
    n_renamed = canonicalize_player_names(data)
    if n_renamed:
        print(f"\nCanonicalized {n_renamed} player name(s) to existing full forms.")

    # Write to Excel
    print("\nAll gates passed — appending to Excel...")
    counts = append_to_excel(data)
    print(f"  Rows added — Game_Log: {counts['Game_Log']} | Batting: {counts['Batting']} | Pitching: {counts['Pitching']} | Fielding: {counts['Fielding']}")

    # Run export.py
    print("\nRunning export.py to regenerate repository.json...")
    result = subprocess.run(
        [sys.executable, str(PIPELINE_DIR / "export.py")],
        capture_output=True, text=True
    )
    if result.returncode == 0:
        print(result.stdout.strip())
    else:
        print(f"export.py failed: {result.stderr}")
        sys.exit(1)

    print("\nIngestion complete.")

    # Audit: final success entry. If earlier log_flag calls already created
    # an entry for this game (e.g. REGISTRY_UNKNOWN_TEAM), this updates
    # ingest_result to "success" without duplicating flags.
    _log_note(game_md_path.stem, ingest_result="success")


if __name__ == "__main__":
    main()
