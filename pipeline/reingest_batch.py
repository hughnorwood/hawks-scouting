#!/usr/bin/env python3
"""Batch re-ingestion tool for audit-failing games.

Operates on Buckets A (focal missing pitcher) and B (focal missing batter) by default,
where Claude non-determinism is the most common root cause and a clean re-ingest usually
resolves the failure. Other buckets require an explicit --force-bucket override.

Workflow per game:
  1. Snapshot all rows for the Game_ID across Batting, Pitching, Fielding, Game_Log
  2. For up to --retries attempts:
       a. Delete the game's rows (removes any partial output from a prior attempt)
       b. Run pipeline/ingest.py on the markdown (cross-checks enabled)
       c. Re-audit via validate.audit_game
       d. On pass → log success, stop. On fail → try again.
  3. After all retries fail → restore original snapshot, log failure, continue.

Safety:
  - Pre-run timestamped backup of data/RiverHill_Repository_Master.xlsx to data/backups/
  - If backup write fails, abort BEFORE touching any game data
  - Per-game snapshot + restore keeps each game atomic
  - Processing order: all A games first, then all B games (never interleaved)

Usage:
  python pipeline/reingest_batch.py                    # all A then all B
  python pipeline/reingest_batch.py --dry-run          # print plan only
  python pipeline/reingest_batch.py --game-id GID      # single game (must be in A or B)
  python pipeline/reingest_batch.py --limit N          # cap at N games
  python pipeline/reingest_batch.py --force-bucket G,D # opt into other buckets
  python pipeline/reingest_batch.py --retries 5        # more attempts per game
"""

import argparse
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

sys.path.insert(0, str(Path(__file__).resolve().parent))
from triage import classify, FOCAL_TEAMS  # noqa: E402
from validate import audit_game  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
EXCEL_FILE = DATA_DIR / "RiverHill_Repository_Master.xlsx"
BACKUP_DIR = DATA_DIR / "backups"
GAMES_DIR = REPO_ROOT / "games"
INGEST_SCRIPT = REPO_ROOT / "pipeline" / "ingest.py"

SHEETS = ["Game_Log", "Batting", "Pitching", "Fielding"]
DEFAULT_BUCKETS = {"A", "B"}


# ─── Backup / restore ────────────────────────────────────────────────────────

def create_full_backup() -> Path:
    """Copy the master Excel file to data/backups/ with a timestamp. Returns the path."""
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d-%H%M%S")
    dest = BACKUP_DIR / f"RiverHill_Repository_Master.{stamp}.xlsx"
    shutil.copy2(EXCEL_FILE, dest)
    if not dest.exists() or dest.stat().st_size != EXCEL_FILE.stat().st_size:
        raise RuntimeError(f"Backup verification failed: {dest}")
    return dest


def snapshot_game_rows(game_id: str) -> dict:
    """Return {sheet_name: [(row_idx, [cell_values])]} for all rows matching Game_ID.

    Row indices are 1-based (openpyxl convention). The header is row 1; data starts at 2.
    """
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=False)
    snapshot = {}
    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            snapshot[sheet] = []
            continue
        ws = wb[sheet]
        matched = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == 1:
                continue  # header
            if row and row[0] == game_id:
                matched.append((idx, list(row)))
        snapshot[sheet] = matched
    wb.close()
    return snapshot


def delete_game_rows(game_id: str) -> dict:
    """Delete every row with this Game_ID from all four sheets. Returns per-sheet counts."""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    counts = {}
    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            counts[sheet] = 0
            continue
        ws = wb[sheet]
        # Collect indices top-down, delete bottom-up so row numbers stay valid
        to_delete = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == 1:
                continue
            if row and row[0] == game_id:
                to_delete.append(idx)
        for idx in reversed(to_delete):
            ws.delete_rows(idx, 1)
        counts[sheet] = len(to_delete)
    wb.save(EXCEL_FILE)
    return counts


def restore_game_rows(snapshot: dict) -> dict:
    """Re-insert rows from a snapshot. Used when re-ingest fails and we need to roll back.

    Strategy: delete any rows currently matching the Game_ID (from the partial ingest),
    then append the original rows back. This preserves the original data, though row
    positions will move to the bottom of each sheet (which is harmless — order is not
    semantic in the repository).
    """
    game_ids = set()
    for rows in snapshot.values():
        for _, values in rows:
            if values and values[0]:
                game_ids.add(values[0])
    for gid in game_ids:
        delete_game_rows(gid)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    counts = {}
    for sheet, rows in snapshot.items():
        if sheet not in wb.sheetnames:
            counts[sheet] = 0
            continue
        ws = wb[sheet]
        for _, values in rows:
            ws.append(values)
        counts[sheet] = len(rows)
    wb.save(EXCEL_FILE)
    return counts


# ─── Worklist construction ───────────────────────────────────────────────────

def build_worklist(allowed_buckets: set, single_game_id: str | None, limit: int | None) -> list:
    """Run audit + triage against current Excel/markdown state; return ordered worklist.

    Ordering: all A games first (sorted by game_id), then all B, then any other
    allowed buckets in alphabetical order.
    """
    if single_game_id:
        md = GAMES_DIR / f"{single_game_id}.md"
        if not md.exists():
            sys.exit(f"Markdown not found: {md}")
        paths = [md]
    else:
        paths = sorted(GAMES_DIR.glob("*.md"))
        paths = [p for p in paths if not p.stem.startswith("UNKNOWN_")]

    results = []
    for p in paths:
        r = audit_game(p)
        if r["status"] != "fail":
            continue
        t = classify(r)
        if t["bucket"] in allowed_buckets:
            results.append(t)

    # Sort: bucket order (A, B, then others alphabetical), then by game_id
    bucket_order = {b: i for i, b in enumerate("ABCDEFGH")}
    results.sort(key=lambda r: (bucket_order.get(r["bucket"], 99), r["game_id"]))

    if limit:
        results = results[:limit]
    return results


# ─── Re-ingest a single game ─────────────────────────────────────────────────

def _run_ingest_once(md_path: Path, log) -> tuple[str, str]:
    """Run ingest.py + re-audit for one attempt. Assumes rows for this game are already
    deleted. Returns (outcome, detail). outcome ∈ {pass, fail, error}.
    Caller is responsible for snapshot/restore — this function does not touch the snapshot.
    On failure, the failed partial ingest (if any) is left in the Excel so the caller can
    decide whether to retry (which will delete them again) or restore."""
    try:
        proc = subprocess.run(
            [sys.executable, str(INGEST_SCRIPT), str(md_path)],
            cwd=str(REPO_ROOT),
            capture_output=True,
            text=True,
            timeout=600,
        )
    except subprocess.TimeoutExpired:
        return "error", "ingest.py timeout (600s)"

    if proc.returncode != 0:
        log(f"    stdout tail: {proc.stdout[-300:].strip()}")
        return "error", f"ingest.py rc={proc.returncode}"

    audit = audit_game(md_path)
    if audit["status"] == "pass":
        return "pass", "audit clean after re-ingest"
    if audit["status"] == "parse_error":
        return "error", f"audit parse_error: {audit.get('error', '?')}"
    discs = audit.get("discrepancies", [])
    detail = "; ".join(
        f"{d['check']} {d['team']}: {str(d.get('details', ''))[:60]}" for d in discs[:3]
    )
    return "fail", detail


def reingest_one(game_id: str, log, retries: int = 3) -> tuple[str, str]:
    """Re-ingest one game with up to `retries` attempts. Returns (outcome, detail).

    On each attempt we snapshot once, delete, call ingest.py, re-audit. If it fails,
    we delete the partial output (if any), try again. Only after all retries fail do we
    restore the original snapshot."""
    md_path = GAMES_DIR / f"{game_id}.md"
    if not md_path.exists():
        return "error", f"markdown missing: {md_path}"

    log(f"  snapshotting existing rows for {game_id}")
    snapshot = snapshot_game_rows(game_id)
    snap_totals = {s: len(rows) for s, rows in snapshot.items()}
    log(f"  snapshot: {snap_totals}")

    last_outcome, last_detail = "error", "no attempts made"
    for attempt in range(1, retries + 1):
        log(f"  attempt {attempt}/{retries}: deleting rows")
        try:
            del_counts = delete_game_rows(game_id)
            log(f"    deleted: {del_counts}")
        except Exception as e:
            last_outcome, last_detail = "error", f"delete failed: {e}"
            break

        log(f"    running ingest.py on {md_path.name}")
        outcome, detail = _run_ingest_once(md_path, log)
        last_outcome, last_detail = outcome, detail

        if outcome == "pass":
            log(f"    ✅ pass on attempt {attempt}")
            return "pass", f"{detail} (attempt {attempt})"

        log(f"    attempt {attempt} → {outcome}: {detail}")
        if outcome == "error" and "parse_error" in detail:
            # parse errors aren't likely to self-correct; bail early
            break

    # All retries exhausted — restore original rows
    log(f"  all {retries} attempts failed; restoring snapshot")
    try:
        restore_game_rows(snapshot)
    except Exception as e:
        log(f"  CRITICAL: restore failed: {e}")
        return "error", f"{last_detail} + RESTORE FAILED: {e}"
    return last_outcome, f"{last_detail} (after {retries} attempts)"


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Batch re-ingest audit-failing games (Buckets A+B by default)"
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Print worklist only; no Excel writes and no API calls")
    parser.add_argument("--game-id", help="Single game id (must be in allowed buckets)")
    parser.add_argument("--limit", type=int, help="Cap the number of games processed")
    parser.add_argument("--force-bucket", default="",
                        help="Comma-separated extra buckets to include (e.g. G,D)")
    parser.add_argument("--retries", type=int, default=3,
                        help="Attempts per game before giving up and restoring (default 3)")
    args = parser.parse_args()

    allowed = set(DEFAULT_BUCKETS)
    if args.force_bucket:
        for b in args.force_bucket.upper().split(","):
            b = b.strip()
            if b:
                allowed.add(b)

    print(f"Allowed buckets: {sorted(allowed)}")
    print(f"Building worklist from audit + triage…")
    worklist = build_worklist(allowed, args.game_id, args.limit)

    if not worklist:
        print("Worklist is empty. Nothing to do.")
        return

    # Print plan
    by_bucket = {}
    for r in worklist:
        by_bucket.setdefault(r["bucket"], []).append(r["game_id"])
    print(f"\nWorklist ({len(worklist)} games):")
    for b in sorted(by_bucket.keys(), key=lambda x: "ABCDEFGH".index(x)):
        print(f"  Bucket {b} — {len(by_bucket[b])} game(s)")
        for gid in by_bucket[b]:
            print(f"    {gid}")

    if args.dry_run:
        print("\n[dry-run] exiting without writes.")
        return

    # Safety backup
    print(f"\nCreating backup…")
    try:
        backup_path = create_full_backup()
        print(f"  backup: {backup_path}")
    except Exception as e:
        sys.exit(f"ABORT: backup failed before any game data was touched: {e}")

    # Log file
    log_path = REPO_ROOT / "pipeline" / "reingest_batch.log"
    log_fh = log_path.open("a")
    stamp = datetime.now().isoformat(timespec="seconds")
    log_fh.write(f"\n=== batch run {stamp} — {len(worklist)} games — backup={backup_path.name} ===\n")

    def log(msg: str):
        print(msg)
        log_fh.write(msg + "\n")
        log_fh.flush()

    # Per-game loop
    results = []
    for i, r in enumerate(worklist, start=1):
        gid = r["game_id"]
        bucket = r["bucket"]
        print(f"\n[{i}/{len(worklist)}] {gid}  (Bucket {bucket})")
        log_fh.write(f"\n[{i}/{len(worklist)}] {gid} (Bucket {bucket})\n")
        try:
            outcome, detail = reingest_one(gid, log, retries=args.retries)
        except Exception as e:
            outcome, detail = "error", f"unhandled exception: {e}"
            log(f"  EXCEPTION: {e}")
        results.append((gid, bucket, outcome, detail))
        print(f"  → {outcome}: {detail}")

    log_fh.close()

    # Summary
    print(f"\n{'='*60}\nRun summary")
    print(f"{'='*60}")
    counts = {"pass": 0, "fail": 0, "error": 0}
    for _, _, o, _ in results:
        counts[o] = counts.get(o, 0) + 1
    print(f"  pass:  {counts['pass']}")
    print(f"  fail:  {counts['fail']}  (original rows restored)")
    print(f"  error: {counts['error']} (original rows restored)")
    print(f"\nPer-game outcomes:")
    for gid, b, o, detail in results:
        print(f"  [{b}] {o:6s} {gid}  {detail[:80]}")

    print(f"\nBackup kept at: {backup_path}")
    print(f"Full log: {log_path}")


if __name__ == "__main__":
    main()
