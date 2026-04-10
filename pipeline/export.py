#!/usr/bin/env python3
"""Export 4 data sheets from the master Excel to public/repository.json."""

import json
import sys
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = REPO_ROOT / "data" / "RiverHill_Repository_Master.xlsx"
JSON_PATH = REPO_ROOT / "public" / "repository.json"

SHEETS = ["Game_Log", "Batting", "Pitching", "Fielding"]
JSON_KEYS = ["gameLog", "batting", "pitching", "fielding"]


def sheet_to_rows(ws):
    """Convert a worksheet to a list of dicts. Numeric fields stay numeric, dates become YYYY-MM-DD strings."""
    rows = []
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        record = {}
        for header, value in zip(headers, row):
            if header is None:
                continue
            if isinstance(value, (date, datetime)):
                record[header] = value.strftime("%Y-%m-%d")
            elif isinstance(value, (int, float)):
                record[header] = int(value) if isinstance(value, float) and value == int(value) else value
            elif value is None:
                record[header] = ""
            else:
                # Try to parse numeric strings
                s = str(value).strip()
                try:
                    record[header] = int(s)
                except ValueError:
                    try:
                        record[header] = float(s)
                    except ValueError:
                        record[header] = s
        rows.append(record)
    return rows


def main():
    if not EXCEL_PATH.exists():
        sys.exit(f"Excel file not found: {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    result = {"exported": datetime.utcnow().isoformat() + "Z"}

    for sheet_name, json_key in zip(SHEETS, JSON_KEYS):
        if sheet_name not in wb.sheetnames:
            sys.exit(f"Sheet '{sheet_name}' not found in workbook")
        result[json_key] = sheet_to_rows(wb[sheet_name])

    wb.close()

    JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(JSON_PATH, "w") as f:
        json.dump(result, f, separators=(",", ":"))

    total = sum(len(result[k]) for k in JSON_KEYS)
    print(f"Exported {total} rows to {JSON_PATH} ({JSON_PATH.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
