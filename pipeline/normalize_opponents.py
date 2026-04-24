"""
normalize_opponents.py  v2

Normalizes the `Opponent` field in the Batting and Pitching sheets of
RiverHill_Repository_Master.xlsx to canonical 4-letter team codes.

Two-pass approach:
  Pass 1 — Mirroring fix
    Rows where Team == Opponent (a non-focal team's rows reflecting their own
    code instead of their actual opponent) are corrected using Game_Log's
    Away_Team / Home_Team columns — no markdown needed.

  Pass 2 — Name normalization
    Full team names (pre-April-15 ingest artifacts) are resolved to canonical
    codes via config.json name_patterns, same logic as ingest.py.

Unresolved values (no pattern match) are reported with counts and example
game_ids so you can open the .md, find the team name, add a config.json
entry, and re-run until the unresolved count reaches zero.

Usage:
    python pipeline/normalize_opponents.py              # dry run
    python pipeline/normalize_opponents.py --apply      # write changes
    python pipeline/normalize_opponents.py --apply --verbose
"""

import json
import shutil
import argparse
from datetime import datetime
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl not found.  Run: pip install openpyxl")

# ─── Paths ────────────────────────────────────────────────────────────────────

REPO_ROOT   = Path(__file__).resolve().parent.parent
EXCEL_PATH  = REPO_ROOT / "data" / "RiverHill_Repository_Master.xlsx"
CONFIG_PATH = REPO_ROOT / "pipeline" / "config.json"
BACKUP_DIR  = REPO_ROOT / "data" / "backups"

DATA_SHEETS = ["Batting", "Pitching"]

# ─── Registry ─────────────────────────────────────────────────────────────────

def load_registry(config_path):
    with open(config_path) as f:
        cfg = json.load(f)

    entries = []   # (pattern_lower, canonical_code)
    codes   = set()

    for team in cfg.get("focal_teams", []) + cfg.get("known_opponents", []):
        code = team["code"]
        codes.add(code)
        for pattern in team.get("name_patterns", []):
            entries.append((pattern.lower(), code))

    # Longest pattern first — same precedence as ingest.py
    entries.sort(key=lambda x: len(x[0]), reverse=True)
    return entries, codes


def resolve_name(raw, entries, known_codes):
    """
    Returns canonical code if resolvable, else None.
    Mirroring is handled separately in Pass 1.
    """
    if not raw or not isinstance(raw, str):
        return None
    raw = raw.strip()
    if raw in known_codes:
        return raw
    raw_lower = raw.lower()
    for pattern, code in entries:
        if pattern in raw_lower:
            return code
    return None


# ─── Game_Log index ───────────────────────────────────────────────────────────

def build_game_index(wb):
    """
    Returns dict: game_id → {"away": away_team_code, "home": home_team_code}
    """
    if "Game_Log" not in wb.sheetnames:
        raise SystemExit("Game_Log sheet not found in workbook.")

    ws      = wb["Game_Log"]
    headers = [c.value for c in ws[1]]

    def idx(name):
        try:
            return headers.index(name)
        except ValueError:
            raise SystemExit(f"Column '{name}' not found in Game_Log.")

    gid_col  = idx("Game_ID")
    away_col = idx("Away_Team")
    home_col = idx("Home_Team")

    index = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        gid  = row[gid_col]
        away = row[away_col]
        home = row[home_col]
        if gid:
            index[str(gid).strip()] = {
                "away": str(away).strip() if away else "",
                "home": str(home).strip() if home else "",
            }
    return index


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply",   action="store_true",
                        help="Write changes to Excel (default: dry run)")
    parser.add_argument("--verbose", action="store_true",
                        help="Print every row change, not just summary")
    args = parser.parse_args()

    if not EXCEL_PATH.exists():
        raise SystemExit(f"Excel file not found: {EXCEL_PATH}")
    if not CONFIG_PATH.exists():
        raise SystemExit(f"config.json not found: {CONFIG_PATH}")

    entries, known_codes = load_registry(CONFIG_PATH)
    print(f"Registry: {len(known_codes)} canonical codes, {len(entries)} name patterns\n")

    if args.apply:
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = BACKUP_DIR / f"RiverHill_Repository_Master_{ts}_pre_normalize.xlsx"
        shutil.copy2(EXCEL_PATH, backup)
        print(f"Backup: {backup}\n")

    wb         = openpyxl.load_workbook(EXCEL_PATH)
    game_index = build_game_index(wb)

    total_mirror = total_norm = total_already = total_unknown = 0
    unknowns = defaultdict(list)   # raw_value → [(sheet, row_idx, game_id)]

    for sheet_name in DATA_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"WARNING: sheet '{sheet_name}' not found, skipping")
            continue

        ws      = wb[sheet_name]
        headers = [c.value for c in ws[1]]

        def col(name):
            try:
                return headers.index(name) + 1
            except ValueError:
                return None

        opp_col  = col("Opponent")
        team_col = col("Team")
        gid_col  = col("Game_ID")

        if None in (opp_col, team_col, gid_col):
            print(f"WARNING: required column missing in {sheet_name}, skipping")
            continue

        s_mirror = s_norm = s_already = s_unknown = 0

        for row_idx in range(2, ws.max_row + 1):
            raw_opp  = ws.cell(row=row_idx, column=opp_col).value
            raw_team = ws.cell(row=row_idx, column=team_col).value
            raw_gid  = ws.cell(row=row_idx, column=gid_col).value

            if raw_opp is None:
                continue

            opp  = str(raw_opp).strip()
            team = str(raw_team).strip() if raw_team else ""
            gid  = str(raw_gid).strip()  if raw_gid  else ""

            # ── Pass 1: Mirroring fix ────────────────────────────────────────
            # Resolve opp to a code first so we can compare against Team.
            # If they match, this row has its own team code in Opponent —
            # the true opponent comes from Game_Log.
            resolved_for_mirror = resolve_name(opp, entries, known_codes) or opp

            if resolved_for_mirror == team and gid in game_index:
                g       = game_index[gid]
                correct = g["home"] if g["away"] == team else g["away"]

                if correct and correct != team:
                    if args.verbose:
                        print(f"  [{sheet_name} r{row_idx}] MIRROR  '{opp}' (team={team}) → '{correct}'")
                    if args.apply:
                        ws.cell(row=row_idx, column=opp_col).value = correct
                    s_mirror += 1
                    continue   # handled — skip Pass 2

            # ── Pass 2: Name normalization ───────────────────────────────────
            resolved = resolve_name(opp, entries, known_codes)

            if resolved is None:
                s_unknown += 1
                unknowns[opp].append((sheet_name, row_idx, gid))
                if args.verbose:
                    print(f"  [{sheet_name} r{row_idx}] UNKNOWN '{opp}'  game={gid}")
            elif resolved == opp:
                s_already += 1
            else:
                if args.verbose:
                    print(f"  [{sheet_name} r{row_idx}] NORM    '{opp}' → '{resolved}'")
                if args.apply:
                    ws.cell(row=row_idx, column=opp_col).value = resolved
                s_norm += 1

        print(f"{sheet_name}:")
        print(f"  Mirroring fixed : {s_mirror}")
        print(f"  Name normalized : {s_norm}")
        print(f"  Already correct : {s_already}")
        print(f"  Unresolved      : {s_unknown}\n")

        total_mirror  += s_mirror
        total_norm    += s_norm
        total_already += s_already
        total_unknown += s_unknown

    # ── Summary ───────────────────────────────────────────────────────────────
    print("── Totals " + "─" * 40)
    print(f"  Mirroring fixed : {total_mirror}")
    print(f"  Name normalized : {total_norm}")
    print(f"  Already correct : {total_already}")
    print(f"  Unresolved      : {total_unknown}")

    if unknowns:
        print("\n── Unresolved values " + "─" * 30)
        print("  (count)  value  →  example game_id\n")
        for val, occurrences in sorted(unknowns.items(), key=lambda x: -len(x[1])):
            example_gid = occurrences[0][2]
            print(f"  ({len(occurrences):3d})  '{val}'  →  {example_gid}")

        print("""
Resolution options:
  A) Readable name  → add to config.json known_opponents:
       { "code": "XXXX", "name_patterns": ["<substring>"] }
  B) Opaque code    → open games/<example_game_id>.md, read the Teams: line
     for the full name, then add as option A.
  C) No .md exists  → data integrity gap; flag for manual review.

Re-run (dry) after updating config.json until unresolved count reaches zero,
then run --apply.
""")

    # ── Write ─────────────────────────────────────────────────────────────────
    if args.apply and (total_mirror + total_norm) > 0:
        wb.save(EXCEL_PATH)
        print(f"Excel saved: {EXCEL_PATH}")
        print("Run export.py next to push a fresh repository.json.")
    elif not args.apply:
        print("\nDRY RUN — no changes written. Re-run with --apply to write.")
    else:
        print("\nNo changes needed.")


if __name__ == "__main__":
    main()
